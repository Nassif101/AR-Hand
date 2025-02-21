#!/usr/bin/env python
# license removed for brevity
import rospy
from geometry_msgs.msg import Quaternion 
from std_msgs.msg import Float64
from openpyxl import Workbook, load_workbook
import os

## I want two files, one for the thumb and one for the palm


wb = Workbook()
ws = wb.active
wb2 = Workbook()
ws2 = wb2.active

def callback(data):
    ws.append([data.x, data.y, data.z, data.w])
    wb.save("data.xlsx")

def _callback(data):
    ws2.append([data.x, data.y, data.z, data.w])
    wb2.save("data1.xlsx")

def main():
    rospy.init_node("listener", anonymous=True)
    rospy.Subscriber("/thumbPIP", Quaternion, callback)
    rospy.Subscriber("/palm_angle", Quaternion, _callback)
    rospy.spin()

if __name__ == "__main__":
    main()